#!/usr/bin/env python3
"""由客戶交付的 Excel 產生 src/data/ 下的資料模組。

來源（TTRI to VJINC/）:
  相關文件資料/資料內容.xlsx   -> 布樣種類 / 組織 / 成分 / 布重 / 布厚 / 彈性 / Pantone
  Asset 檔案/Asset.xlsx        -> 組織（布樣清單）/ 背景 / 模型

用法: python3 scripts/build-fabric-data.py
"""

import json
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DELIVERY = os.path.join(ROOT, 'TTRI to VJINC')
OUT_DIR = os.path.join(ROOT, 'src', 'data')

CONTENT_XLSX = os.path.join(DELIVERY, '相關文件資料', '資料內容.xlsx')
ASSET_XLSX = os.path.join(DELIVERY, 'Asset 檔案', 'Asset.xlsx')
SWATCH_DIR = os.path.join(DELIVERY, 'Asset 檔案', '組織瀏覽圖')
WEAVE_DIR = os.path.join(DELIVERY, 'Asset 檔案', '組織')

# 模型公分尺寸 -> 貼圖 tiling 的換算基準（沿用既有 data.js 的比例：tilingX = 公分 / 9.144，tilingY = tilingX * 0.8）
TILING_CM_PER_REPEAT = 9.144
TILING_Y_RATIO = 0.8

# 布種代碼。LF（貼合）為 Specs2VS 支援但本次未交付素材。
FABRIC_CATEGORIES = [
    {'value': 'WO', 'text': '梭織', 'assetPrefix': 'WO1', 'legacyPrefix': '1'},
    {'value': 'CW', 'text': '圓編', 'assetPrefix': 'CW2', 'legacyPrefix': '2'},
    {'value': 'WK', 'text': '經編', 'assetPrefix': 'WK3', 'legacyPrefix': '3'},
    {'value': 'LF', 'text': '貼合', 'assetPrefix': None, 'legacyPrefix': None},
]

# 組織名稱 -> 代碼。代碼沿用既有前端定義；泡泡布為本次新增（資料內容.xlsx 有列但未給代碼）。
WEAVE_CODES = {
    'CW': [
        ('CWD020', '雙面PK布'),
        ('CWD050', '雙面羅紋'),
        ('CWD070', '雙面洞洞布'),
        ('CWS010', '單面平紋'),
        ('CWS060', '單面刷毛布'),
        ('CWS070', '單面組織布'),
        ('CWS080', '單面洞洞布'),
    ],
    'WO': [
        ('WOS010', '平紋'),
        ('WOS020', '斜紋'),
        ('WOS030', '緞紋'),
        ('WOS040', '泡泡布'),  # 代碼為暫定值，待客戶確認
    ],
    'WK': [
        ('WKS010', '經編平紋'),
        ('WKS020', '經編網布'),
    ],
}

PROVISIONAL_WEAVE_CODES = {'WOS040'}

STRETCH_OPTIONS = [
    ('NA', '無彈性'),
    ('warp stretch', '經向彈性'),
    ('weft stretch', '緯向彈性'),
    ('4 way stretch', '四向彈性'),
]

# Pattern 生成 API（Pattern_API.xlsx / Schema_Generate_pattern）
PATTERN_TASKS = ['seamless', 'tiling', 'general']
PATTERN_STYLES = ['', 'Abstract', 'Gouache', 'Impressionistic', 'Line Art', 'Low-poly',
                  'Minimalist', 'Pastel', 'Sketch', 'Vector Art', 'Watercolor']
PATTERN_SERVERS = [
    {'value': 'Gemini', 'models': ['Gemini2.5', 'Gemini-3', 'Imagen'], 'default': 'Gemini2.5'},
    {'value': 'TTRI', 'models': ['dev', 'schnell'], 'default': 'dev'},
]
PATTERN_IMAGE_SIZES = [512, 1024]


def norm(text):
    return re.sub(r'\s+', '', str(text or ''))


def js(value, indent=0):
    """輸出可讀性較好的 JS literal。"""
    return json.dumps(value, ensure_ascii=False, indent=indent or None)


def header(source):
    return (f'// 本檔由 scripts/build-fabric-data.py 依 {source} 自動產生，請勿手動編輯。\n')


def load_content():
    wb = openpyxl.load_workbook(CONTENT_XLSX, data_only=True)

    compositions = []
    for row in wb['成分'].iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        compositions.append({
            'value': row[1],
            'text': row[3],
            'name': row[2],
            'group': row[4],
        })

    def parse_range(sheet):
        text = str(wb[sheet].cell(row=2, column=1).value or '')
        nums = [float(n) for n in re.findall(r'\d+(?:\.\d+)?', text)]
        unit = str(wb[sheet].cell(row=2, column=2).value or '').replace('單位為', '').strip()
        return {'min': nums[0], 'max': nums[1], 'unit': unit}

    return {
        'compositions': compositions,
        'gsm': parse_range('布重'),
        'thickness': parse_range('布厚'),
    }


def load_assets():
    wb = openpyxl.load_workbook(ASSET_XLSX, data_only=True)

    # 組織：編號 / 布樣種類 / 組織 / 成分（成分為多列合併儲存格）
    fabrics = []
    for row in wb['組織'].iter_rows(min_row=2, values_only=True):
        code, category_text, weave_text, comp = row[0], row[1], row[2], row[3]
        if code and str(code).strip():
            fabrics.append({
                'code': str(code).strip(),
                'categoryText': category_text,
                'weaveText': norm(weave_text),
                'compositions': [],
            })
        if comp and fabrics:
            name, _, percent = str(comp).partition('：')
            fabrics[-1]['compositions'].append({
                'name': name.strip(),
                'percent': float(re.sub(r'[^\d.]', '', percent) or 0),
            })

    backgrounds = []
    for row in wb['背景'].iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        backgrounds.append({
            'index': int(row[0]),
            'kind': row[1],
            'text': row[2],
            'file': None if not row[3] or row[3] == '無' else row[3],
        })

    models = []
    for row in wb['模型'].iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        models.append({
            'index': int(row[0]),
            'text': row[1],
            'group': row[2],
            'file': row[3],
            'sizeX': float(row[4]),
            'sizeY': float(row[5]),
            'public': str(row[6] or '').upper() == 'Y',
        })

    return {'fabrics': fabrics, 'backgrounds': backgrounds, 'models': models}


def write_options(content):
    weave_by_type = {}
    for cat, items in WEAVE_CODES.items():
        weave_by_type[cat] = [
            {'value': code, 'text': text, **({'provisional': True} if code in PROVISIONAL_WEAVE_CODES else {})}
            for code, text in items
        ]

    lines = [header('相關文件資料/資料內容.xlsx 與 Pattern 生成API 資料/Pattern_API.xlsx'), '']

    lines.append('// Specs2VS 支援的布種。LF（貼合）本次未交付素材，僅保留代碼。')
    lines.append(f'export const fabricCategories = {js([{k: v for k, v in c.items() if k in ("value", "text")} for c in FABRIC_CATEGORIES], 2)};\n')
    lines.append('// 交付素材涵蓋的布種（供 UI 下拉使用）')
    lines.append("export const fabricTypes = fabricCategories.filter((item) => item.value !== 'LF');\n")

    lines.append('// 組織選項。標記 provisional 者為資料內容.xlsx 有列出但客戶尚未給代碼。')
    lines.append(f'export const fabricationsByType = {js(weave_by_type, 2)};\n')

    lines.append(f'export const compositions = {js(content["compositions"], 2)};\n')

    lines.append(f'export const stretchOptions = {js([{"value": v, "text": t} for v, t in STRETCH_OPTIONS], 2)};\n')

    limits = {
        'gsm': content['gsm'],
        'thickness': content['thickness'],
        # Schema_Specs2VS 宣告的模型有效範圍，比資料內容.xlsx 窄；超出時提示可信度較低。
        'apiGsm': {'min': 50, 'max': 400, 'unit': 'g/m2'},
        'apiThickness': {'min': 0.05, 'max': 1.5, 'unit': 'mm'},
        'opPercentage': {'min': 0, 'max': 60, 'unit': '%'},
    }
    lines.append('// gsm / thickness 為資料內容.xlsx 的輸入限制；api* 為 Specs2VS 模型的有效範圍。')
    lines.append(f'export const specLimits = {js(limits, 2)};\n')

    lines.append('// Pattern 生成 API（Schema_Generate_pattern）')
    lines.append(f'export const patternTasks = {js(PATTERN_TASKS, 2)};')
    lines.append(f'export const patternStyles = {js(PATTERN_STYLES, 2)};')
    lines.append(f'export const patternServers = {js(PATTERN_SERVERS, 2)};')
    lines.append(f'export const patternImageSizes = {js(PATTERN_IMAGE_SIZES)};\n')

    lines.append('export const compositionByCode = Object.fromEntries(compositions.map((item) => [item.value, item]));')
    lines.append('export const fabricCategoryByCode = Object.fromEntries(fabricCategories.map((item) => [item.value, item]));\n')

    path = os.path.join(OUT_DIR, 'fabric-options.js')
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(lines))
    return path, len(content['compositions'])


def write_library(content, assets):
    comp_by_name = {c['text']: c for c in content['compositions']}
    weave_code_by_text = {}
    for cat, items in WEAVE_CODES.items():
        for code, text in items:
            weave_code_by_text[(cat, norm(text))] = code
    cat_by_text = {c['text']: c for c in FABRIC_CATEGORIES}
    prefix_to_cat = {c['assetPrefix']: c['value'] for c in FABRIC_CATEGORIES if c['assetPrefix']}

    swatches = {f[:-6] for f in os.listdir(SWATCH_DIR) if f.endswith('_G.png')}
    weave_dirs = {d for d in os.listdir(WEAVE_DIR) if os.path.isdir(os.path.join(WEAVE_DIR, d))}

    # 交付的 u3m 資料夾以「舊制編號」命名（3-0000001），對應 WK3-0000001
    legacy_by_code = {}
    for c in FABRIC_CATEGORIES:
        if not c['assetPrefix']:
            continue
        for d in weave_dirs:
            prefix, _, serial = d.partition('-')
            if prefix == c['legacyPrefix']:
                legacy_by_code[f'{c["assetPrefix"]}-{serial}'] = d

    items = []
    for fabric in assets['fabrics']:
        code = fabric['code']
        category = prefix_to_cat.get(code.split('-')[0])
        entry = {
            'code': code,
            'categoryCode': category,
            'categoryText': fabric['categoryText'],
            'weaveCode': weave_code_by_text.get((category, fabric['weaveText'])),
            'weaveText': fabric['weaveText'],
            'compositions': [
                {
                    'code': comp_by_name.get(c['name'], {}).get('value'),
                    'text': c['name'],
                    'percent': c['percent'],
                }
                for c in fabric['compositions']
            ],
        }
        entry['compositionText'] = '、'.join(
            f'{_fmt(c["percent"])}% {c["name"]}' for c in fabric['compositions']
        )
        entry['swatch'] = f'/assets/swatches/{code}_G.png' if code in swatches else None

        legacy = legacy_by_code.get(code)
        if legacy:
            src_dir = os.path.join(WEAVE_DIR, legacy)
            physics = next((f for f in os.listdir(src_dir) if f.endswith('.json')), None)
            entry['legacyCode'] = legacy
            entry['u3m'] = {
                'dir': f'/assets/fabrics/{legacy}',
                'file': f'/assets/fabrics/{legacy}/{legacy}.u3m',
                'physics': f'/assets/fabrics/{legacy}/{physics}' if physics else None,
                'preview': f'/assets/fabrics/{legacy}/{legacy}.png',
                'textures': {
                    key.lower(): f'/assets/fabrics/{legacy}/textures/{legacy}_{key}.jpg'
                    for key in ('BASE', 'NRM', 'ROUGH', 'DISP', 'MTL', 'ALPHA')
                },
                'renders': {
                    label: f'/assets/fabrics/{legacy}/{label}.jpg'
                    for label in ('beauty', 'crumpled_cloth', 'folded_cloth')
                },
            }
        else:
            entry['u3m'] = None
        items.append(entry)

    orphan_swatches = sorted(swatches - {f['code'] for f in assets['fabrics']})

    lines = [header('Asset 檔案/Asset.xlsx（組織）與 Asset 檔案/組織、組織瀏覽圖'), '']
    lines.append(f'export const fabricLibrary = {js(items, 2)};\n')
    lines.append('// 有瀏覽圖但未列在 Asset.xlsx 的編號（待客戶確認）')
    lines.append(f'export const unlistedSwatches = {js(orphan_swatches)};\n')
    lines.append('export const fabricByCode = Object.fromEntries(fabricLibrary.map((item) => [item.code, item]));')
    lines.append('export const fabricsWithU3M = fabricLibrary.filter((item) => item.u3m);\n')

    path = os.path.join(OUT_DIR, 'fabric-library.js')
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(lines))
    return path, len(items), sum(1 for i in items if i['u3m']), orphan_swatches


def write_preview_assets(assets):
    models = []
    for model in assets['models']:
        tiling_x = round(model['sizeX'] / TILING_CM_PER_REPEAT, 2)
        models.append({
            'value': model['index'],
            'displayname': model['text'],
            'group': model['group'],
            'file': f'/assets/models/{model["file"]}',
            'modelType': os.path.splitext(model['file'])[1].lstrip('.'),
            'sizeX': model['sizeX'],
            'sizeY': model['sizeY'],
            'tilingX': tiling_x,
            'tilingY': round(tiling_x * TILING_Y_RATIO, 2),
            'public': model['public'],
        })

    backgrounds = []
    for bg in assets['backgrounds']:
        backgrounds.append({
            'value': bg['index'],
            'displayname': bg['text'],
            'kind': bg['kind'],
            # 素色背景不使用 HDR，改以顏色填充（Asset.xlsx 備註）
            'file': f'/assets/hdr/{bg["file"]}' if bg['file'] else '#e5e5e5',
            'intensity': 0.6 if bg['kind'] != 'HDR' else (0.3 if 'studio' in (bg['file'] or '') else 0.6),
        })

    lines = [header('Asset 檔案/Asset.xlsx（模型、背景）'), '']
    lines.append(f'// tilingX = 模型公分尺寸 / {TILING_CM_PER_REPEAT}，tilingY = tilingX * {TILING_Y_RATIO}')
    lines.append(f'export const previewModels = {js(models, 2)};\n')
    lines.append(f'export const hdriOptions = {js(backgrounds, 2)};\n')
    lines.append('export const previewModelByValue = Object.fromEntries(previewModels.map((item) => [item.value, item]));\n')

    path = os.path.join(OUT_DIR, 'preview-assets.js')
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write('\n'.join(lines))
    return path, len(models), len(backgrounds)


def _fmt(value):
    return int(value) if float(value).is_integer() else value


def main():
    if not os.path.isdir(DELIVERY):
        sys.exit(f'找不到交付資料夾: {DELIVERY}')

    content = load_content()
    assets = load_assets()

    p1, n_comp = write_options(content)
    p2, n_fabric, n_u3m, orphans = write_library(content, assets)
    p3, n_model, n_bg = write_preview_assets(assets)

    print(f'{os.path.relpath(p1, ROOT)}  成分 {n_comp} 筆')
    print(f'{os.path.relpath(p2, ROOT)}  布樣 {n_fabric} 筆（{n_u3m} 筆有 U3M），未列清單的瀏覽圖 {orphans}')
    print(f'{os.path.relpath(p3, ROOT)}  模型 {n_model} 個、背景 {n_bg} 個')


if __name__ == '__main__':
    main()
